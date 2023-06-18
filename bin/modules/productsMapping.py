import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
from pathlib import Path

class ProductDataMapper:
    def __init__(self, fileLoc: str, products: pd.DataFrame) -> None:
        self.fileLoc = fileLoc
        self.dfProduct = products
        self.dfProductIndexSet = set(self.dfProduct.index)
        self.workbook = openpyxl.load_workbook(self.fileLoc)
        self.includeNaLast = True
        self.settings = pd.DataFrame(columns=(['Check', 'Sheet', 'NUM'] + self.dfProduct.columns.to_list()), data={'Check': True, 'Sheet': self.workbook.sheetnames})
        self.processed = False
    
    def __del__(self):
        if self.workbook is not None:
            self.workbook.close()

    def isSettingsSet(self):
        return self.settings['Check'].any() and (self.includeNaLast or self.settings.iloc[:, 2:].any().any())

    def mapToExcelSheet(self, sheet: str, idCol: str, colMap: dict, includeNaLast: bool=False):
        worksheet = self.workbook[sheet]
        colMapSorted = dict(sorted(colMap.items(), key=lambda x: (x[1] is None, x[1])))
        for csvCol, xlsxCol in colMapSorted.items():
            if pd.isna(xlsxCol):
                if includeNaLast:
                    i = 1
                    while (worksheet.max_column + i) in colMap.values():
                        i += 1
                    colMap[csvCol] = worksheet.max_column + i
                else:
                    colMap.pop(csvCol)
            else:
                colMap[csvCol] = column_index_from_string(xlsxCol)
        if colMap:
            if pd.isna(idCol) or idCol == '':
                for row in worksheet.iter_rows():
                    for cell in row:
                        num = str(cell.value)
                        if num in self.dfProductIndexSet:
                            for csvCol, xlsxCol in colMap.items():
                                worksheet.cell(row=cell.row, column=xlsxCol).value = self.dfProduct[csvCol][num]
                            continue
            else:
                for cell in worksheet[idCol]:
                    num = str(cell.value)
                    if num in self.dfProductIndexSet:
                        row = self.dfProduct.loc[num]
                        for csvCol, xlsxCol in colMap.items():
                                worksheet.cell(row=cell.row, column=xlsxCol).value = row[csvCol]

    def mapToExcel(self, save: bool=False, saveFolder: str=None) -> bool:
        if self.isSettingsSet():
            for _, row in self.settings.iterrows():
                self.mapToExcelSheet(sheet=row['Sheet'], idCol=row['NUM'], colMap=row.iloc[3:].to_dict(), includeNaLast=self.includeNaLast)
                self.processed = True
            if save:
                self.save(saveFolder)
            return True
        else:
            return False

    def save(self, saveFolder: str=None):
        if saveFolder is None:
            self.workbook.save(self.fileLoc)
        else:
            fileLoc = Path(self.fileLoc)
            self.workbook.save(Path(saveFolder) / f"{fileLoc.stem}_output{fileLoc.suffix}")